from socket import *
import json
import openpyxl
import os

class friendlist:
    filename = 'C://Users/Jihui/Documents/GitHub/CS322Project/chatProject/server/friendList.txt'
    list_friend = {}
    def __init__(self):
        pass

    # def add_friend(self, user, friend):
        # if friend not in self.list_friend:
        #
        #     json.dump(self.all_name, open(self.filename, 'w'))
        #if os.path.exists(os.path.join(os.getcwd()+r'\friendList.xls')):
                # pass
            # recreate part
            # if False:
            #     os.remove(path)
            #     wk = openpyxl.Workbook()
            #     sheet = wk.get_sheet_by_name('sheet')
            #     add_sheet = []
            #     sheet.title = add_sheet[0]
            #     for i in range(1,len(add_sheet)):
            #         wk.create_sheet(add_sheet[i])
            #     sheet_num = wk.get_sheet_names()
            #     last_sheet = len(sheet_num) - 1
            #     sheet_index = sheet_num.index(sheet_num[last_sheet])
            #     wk.active = sheet_index
            #     wk.save(path)
            #     print("friendList.xlsx had been deleted and recreated")
            # else:
            #     pass
        # else:
        #     wb = openpyxl.Workbook("utf-8")
        #     wb.get_sheet_names()
        #     sheet = wb.active
        #     sheet.title = 'FriendList'
            # print("Create successful!")

    def addFriend(self, username, file):
        pass
        # dic = {}
        # with open('friendlist.txt', "r") as file:
        #     all_lines = file.readlines()
        # file.close()
        # for line in all_lines:
        #     for data in line:
        #         a = data.split()
        #         a[0] = username
        #         dic[a[0]] = a[1]



    def removeFriend(self):
        pass

    def display(self):
        pass
